from __future__ import annotations

import csv
import io
import re
from pathlib import Path

from word_assistance.safety.policies import sanitize_untrusted_text
from word_assistance.services.llm import LLMService

WORD_RE = re.compile(r"[A-Za-z0-9][A-Za-z0-9'-]{1,31}")

PHRASAL_PARTICLES = {
    "up",
    "down",
    "in",
    "out",
    "off",
    "on",
    "away",
    "over",
    "around",
    "through",
    "across",
}

IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".heic", ".bmp", ".webp"}
OCR_STRENGTHS = {"FAST", "BALANCED", "ACCURATE"}
IMPORT_HEADER_HINT_WORDS = {
    "north",
    "shore",
    "coaching",
    "college",
    "develop",
    "your",
    "english",
    "skills",
    "level",
    "lesson",
    "page",
    "spelling",
    "list",
    "word",
    "words",
    "definitions",
    "weekly",
    "website",
    "student",
    "grouping",
    "hear",
    "each",
    "said",
}
IMPORT_DEFINITION_LINKERS = {
    "the",
    "a",
    "an",
    "to",
    "in",
    "of",
    "for",
    "with",
    "on",
    "by",
    "where",
    "who",
    "that",
    "being",
}
IMPORT_NOISE_WORDS = {
    "north",
    "shore",
    "develop",
    "your",
    "english",
    "skills",
    "lesson",
    "level",
    "page",
    "spelling",
    "list",
    "word",
    "words",
    "definition",
    "definitions",
    "weekly",
    "website",
    "student",
    "grouping",
    "hear",
    "here",
    "each",
    "also",
    "using",
    "used",
    "log",
    "said",
    "see",
}


def extract_text_from_bytes(filename: str, payload: bytes, *, ocr_strength: str = "BALANCED") -> str:
    suffix = Path(filename).suffix.lower()
    if suffix in {".txt", ".md", ".log"}:
        return payload.decode("utf-8", errors="ignore")
    if suffix == ".csv":
        return _extract_from_csv(payload)
    if suffix in {".xlsx", ".xlsm", ".xls"}:
        return _extract_from_excel(payload)
    if suffix == ".pdf":
        return _extract_from_pdf(payload)
    if suffix in IMAGE_SUFFIXES:
        return _extract_from_image(payload=payload, suffix=suffix, ocr_strength=ocr_strength)
    return payload.decode("utf-8", errors="ignore")


def _extract_from_csv(payload: bytes) -> str:
    data = payload.decode("utf-8", errors="ignore")
    buf = io.StringIO(data)
    reader = csv.reader(buf)
    values: list[str] = []
    for row in reader:
        values.extend(cell.strip() for cell in row if cell.strip())
    return "\n".join(values)


def _extract_from_excel(payload: bytes) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("Excel 解析需要安装 openpyxl") from exc

    wb = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    chunks: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if isinstance(cell, str) and cell.strip():
                    chunks.append(cell.strip())
    return "\n".join(chunks)


def _extract_from_pdf(payload: bytes) -> str:
    try:
        from pypdf import PdfReader
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("PDF 解析需要安装 pypdf") from exc

    reader = PdfReader(io.BytesIO(payload))
    chunks: list[str] = []
    for page in reader.pages:
        text = page.extract_text() or ""
        if text.strip():
            chunks.append(text.strip())
    return "\n".join(chunks)


def _extract_from_image(payload: bytes, suffix: str, ocr_strength: str = "BALANCED") -> str:
    strength = _normalize_ocr_strength(ocr_strength)
    ocr_outputs: list[str] = []

    try:
        from PIL import Image, ImageEnhance, ImageFilter, ImageOps

        image = Image.open(io.BytesIO(payload)).convert("RGB")
        variants = _build_ocr_variants(
            image=image,
            image_enhance=ImageEnhance,
            image_filter=ImageFilter,
            image_ops=ImageOps,
            strength=strength,
        )
        configs = _ocr_psm_configs(strength)

        try:
            import pytesseract

            for variant in variants:
                for config in configs:
                    text = pytesseract.image_to_string(
                        variant,
                        lang="eng",
                        config=config,
                    )
                    text = text.strip()
                    if text:
                        ocr_outputs.append(text)
        except Exception:
            pass
    except Exception:
        pass

    best = _pick_best_ocr_text(ocr_outputs)
    best_score = _score_ocr_text(best)
    best_tokens = best_score[0]
    should_try_llm = strength == "ACCURATE" or (strength == "BALANCED" and best_tokens < 3)

    if not should_try_llm:
        return best

    mime = _suffix_to_mime(suffix)
    llm_text = LLMService().ocr_from_image_bytes(payload=payload, mime_type=mime)
    llm_text = llm_text.strip()
    if not llm_text:
        return best

    llm_score = _score_ocr_text(llm_text)
    if llm_score >= best_score:
        return llm_text
    if strength == "ACCURATE" and llm_score[0] >= max(2, best_tokens):
        return llm_text
    return best


def _pick_best_ocr_text(candidates: list[str]) -> str:
    if not candidates:
        return ""

    unique_candidates: list[str] = []
    seen: set[str] = set()
    for text in candidates:
        norm = " ".join(extract_normalized_tokens(text))
        if norm and norm in seen:
            continue
        if norm:
            seen.add(norm)
        unique_candidates.append(text)
    return max(unique_candidates, key=_score_ocr_text)


def _score_ocr_text(text: str) -> tuple[int, int, int]:
    tokens = extract_normalized_tokens(text)
    alpha_count = sum(1 for ch in text if ch.isalpha())
    digit_count = sum(1 for ch in text if ch.isdigit())
    alpha_ratio = int(alpha_count * 100 / max(1, len(text)))
    return (len(tokens), alpha_ratio, -digit_count)


def _build_ocr_variants(image, image_enhance, image_filter, image_ops, strength: str) -> list:
    variants = [image]
    gray = image_ops.grayscale(image)
    variants.append(gray)
    variants.append(image_ops.autocontrast(gray))

    if strength in {"BALANCED", "ACCURATE"}:
        sharpened = image_enhance.Sharpness(gray).enhance(2.2)
        variants.append(sharpened)
        thresholded = sharpened.point(lambda x: 255 if x > 145 else 0)
        variants.append(thresholded)
        variants.append(thresholded.filter(image_filter.MedianFilter(size=3)))

    if strength == "ACCURATE":
        denoised = gray.filter(image_filter.MedianFilter(size=3))
        variants.append(denoised)
        variants.append(image_ops.autocontrast(denoised))

    return variants


def _ocr_psm_configs(strength: str) -> list[str]:
    if strength == "FAST":
        return ["--oem 3 --psm 6"]
    if strength == "ACCURATE":
        return ["--oem 3 --psm 6", "--oem 3 --psm 11", "--oem 3 --psm 4", "--oem 3 --psm 3"]
    return ["--oem 3 --psm 6", "--oem 3 --psm 11"]


def _normalize_ocr_strength(value: str) -> str:
    strength = str(value or "BALANCED").strip().upper()
    if strength not in OCR_STRENGTHS:
        return "BALANCED"
    return strength


def _suffix_to_mime(suffix: str) -> str:
    return {
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".heic": "image/heic",
        ".bmp": "image/bmp",
        ".webp": "image/webp",
    }.get(suffix, "image/png")


def extract_candidates(text: str) -> list[str]:
    cleaned = sanitize_untrusted_text(text)
    tokens = extract_normalized_tokens(cleaned)

    candidates: list[str] = []
    seen: set[str] = set()

    for idx, token in enumerate(tokens):
        lemma = simple_lemma(token)
        if lemma not in seen:
            seen.add(lemma)
            candidates.append(lemma)

        if idx + 1 < len(tokens) and tokens[idx + 1] in PHRASAL_PARTICLES:
            phrase = f"{lemma} {tokens[idx + 1]}"
            if phrase not in seen:
                seen.add(phrase)
                candidates.append(phrase)

    return candidates


def extract_document_vocab_candidates(text: str, *, max_words: int = 300) -> list[str]:
    cleaned = sanitize_untrusted_text(text)
    lines = [line.strip() for line in cleaned.splitlines() if line.strip()]
    collected: list[str] = []
    seen: set[str] = set()

    for line in lines:
        tokens = [normalize_word(tok) for tok in WORD_RE.findall(line)]
        tokens = [tok for tok in tokens if len(tok) >= 2]
        if not tokens:
            continue
        if _looks_like_import_header(tokens):
            continue

        for token in _extract_candidates_from_line(tokens):
            lemma = simple_lemma(token)
            if not _is_importable_vocab(lemma) or lemma in seen:
                continue
            seen.add(lemma)
            collected.append(lemma)
            if len(collected) >= max_words:
                return collected

    if collected:
        return collected

    # Fall back to generic extraction when line-level parsing does not yield enough words.
    for token in extract_candidates(cleaned):
        if not _is_importable_vocab(token) or token in seen:
            continue
        seen.add(token)
        collected.append(token)
        if len(collected) >= max_words:
            break

    return collected


def extract_normalized_tokens(text: str) -> list[str]:
    tokens = [normalize_word(tok) for tok in WORD_RE.findall(text)]
    return [tok for tok in tokens if len(tok) >= 2]


def normalize_word(word: str) -> str:
    return word.strip("'\".,;:!?()[]{}<>").lower()


def simple_lemma(word: str) -> str:
    if " " in word:
        parts = word.split(" ")
        return " ".join(simple_lemma(part) for part in parts)

    if len(word) > 5 and word.endswith("ies"):
        return word[:-3] + "y"
    if len(word) > 4 and word.endswith("ing"):
        root = word[:-3]
        if len(root) > 2 and root[-1] == root[-2]:
            root = root[:-1]
        if root.endswith("v"):
            root = root + "e"
        return root
    if len(word) > 3 and word.endswith("ed"):
        root = word[:-2]
        if len(root) > 2 and root[-1] == root[-2]:
            root = root[:-1]
        return root
    if len(word) > 3 and word.endswith("es") and not word.endswith(("ses", "xes")):
        return word[:-2]
    # Keep lexical endings that are often part of base forms, not plural suffixes.
    if len(word) > 3 and word.endswith(("ous", "us", "is", "ss")):
        return word
    if len(word) > 3 and word.endswith("s") and not word.endswith("ss"):
        return word[:-1]
    return word


def _extract_candidates_from_line(tokens: list[str]) -> list[str]:
    if len(tokens) == 1:
        return tokens

    if _looks_like_definition_row(tokens):
        return [tokens[0]]

    if len(tokens) <= 3 and all(_is_importable_vocab(tok) for tok in tokens):
        return tokens

    return []


def _looks_like_definition_row(tokens: list[str]) -> bool:
    if len(tokens) < 2:
        return False
    context_window = tokens[1:5]
    if any(tok in IMPORT_DEFINITION_LINKERS for tok in context_window):
        return True
    return False


def _looks_like_import_header(tokens: list[str]) -> bool:
    if len(tokens) < 2:
        return False
    hint_hits = sum(1 for tok in tokens if tok in IMPORT_HEADER_HINT_WORDS)
    if hint_hits >= 2:
        return True
    if tokens[0] in {"lesson", "level", "page", "spelling", "definitions"}:
        return True
    return False


def _is_importable_vocab(token: str) -> bool:
    if not token:
        return False
    if not re.fullmatch(r"[a-z][a-z'-]{1,32}", token):
        return False
    if token in IMPORT_NOISE_WORDS:
        return False
    return True
